from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

from sqlalchemy import func

from app.models.inventory import Inventory
from app.models.item import Item
from app.core.constants import CATEGORY_LABELS


def generate_excel_report(db, center, is_global=False):

    wb = Workbook()
    ws = wb.active
    ws.title = "Donaciones"

    # =========================
    # LOGO
    # =========================
    try:
        logo = Image("app/static/Logo El Sistema.jpg")
        logo.width = 120
        logo.height = 120
        ws.add_image(logo, "B1")
    except:
        pass

    # =========================
    # QUERY DATA
    # =========================
    if is_global:

        rows = (
            db.query(
                Item,
                func.coalesce(func.sum(Inventory.quantity), 0)
            )
            .join(Inventory, Inventory.item_id == Item.id)
            .group_by(Item.id)
            .order_by(Item.name)
            .all()
        )

        title = "Recaudación consolidada - Gerencia Estadal Bolívar"

    else:

        rows = (
            db.query(
                Item,
                func.coalesce(Inventory.quantity, 0)
            )
            .join(Inventory, Inventory.item_id == Item.id)
            .filter(Inventory.center_id == center.id)
            .order_by(Item.name)
            .all()
        )

        title = center.name

    # =========================
    # TITLE
    # =========================
    ws.merge_cells("A8:C8")

    cell = ws["A8"]
    cell.value = title
    cell.font = Font(size=18, bold=True)
    cell.alignment = Alignment(horizontal="center")

    # =========================
    # HEADERS
    # =========================
    start = 10

    headers = [
        "Donación",
        "Categoría",
        "Unidades recaudadas"
    ]

    for col, value in enumerate(headers, start=1):
        c = ws.cell(row=start, column=col)
        c.value = value
        c.font = Font(bold=True)

    # =========================
    # DATA
    # =========================
    row_num = start + 1

    for item, qty in rows:

        category_key = str(item.category).split(".")[-1].lower()

        category = CATEGORY_LABELS.get(category_key, category_key)

        ws.cell(row=row_num, column=1).value = item.name
        ws.cell(row=row_num, column=2).value = category
        ws.cell(row=row_num, column=3).value = int(qty or 0)

        row_num += 1

    # =========================
    # COLUMN WIDTHS
    # =========================
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    # =========================
    # SIGNATURE
    # =========================
    row_num += 4

    ws.merge_cells(f"A{row_num}:C{row_num}")

    ws.cell(row=row_num, column=1).value = "Firma de la Gerencia"
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")

    # =========================
    # OUTPUT
    # =========================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output